# coding:utf-8

import pymysql
import openpyxl
import time
# # db = pymysql.connect(host='127.0.0.1', user='root',  password='123456', db='demo')
# db = pymysql.connect(host='rm-2zekq87i0gvxaxvph.mysql.rds.aliyuncs.com', user='pointcard_ro', password='oc1_eqbVqgLjo0', db='pointcard')
# cursor = db.cursor()
# sql = 'select card_name from tb_point_card;'
# cursor.execute(sql)
# data = cursor.fetchall()
# print(data[0])
# db.close()
#

# db2 = pymysql.connect(host='rm-2ze9ycl21k1slu15zrw.mysql.rds.aliyuncs.com', user='english_radio_rw', password='english_radio_RW', db='english_radio')
# db = pymysql.connect(host='rm-2zew70125i56c8002rw.mysql.rds.aliyuncs.com', user='library_rw',
#                      password='c6BVh0p9NEMbbjOp', db='chinese_library')
# cursor2 = db2.cursor()
# sql = 'SELECT id from tb_radio_listenlist LIMIT 10000;'
# sql2 = 'select b_id from tb_library_book limit 10000;'
# cursor2.execute(sql)
# data = cursor2.fetchall()
# print(data)
# import openpyxl
# wb = openpyxl.Workbook()
# ws = wb.active
# for i in range(len(data)):
#     doc='A'+str(i+1)
#     print(doc)
#     ws[doc] = str(data[i][0])
# wb.save('temp_listenId.xlsx')


# import pymysql
# # db = pymysql.connect(host='127.0.0.1', user='root',  password='123456', db='demo')
#
# db2 = pymysql.connect(host='rm-2ze9ycl21k1slu15zrw.mysql.rds.aliyuncs.com', user='english_radio_rw', password='english_radio_RW', db='english_radio')
# cursor2 = db2.cursor()
# sql2 = 'SELECT id from tb_radio_listenlist LIMIT 10000;'
# cursor2.execute(sql2)
# data2 = cursor2.fetchall()
# print(data2)
# db = pymysql.connect(host='rm-2zew70125i56c8002rw.mysql.rds.aliyuncs.com', user='library_rw', password='c6BVh0p9NEMbbjOp', db='chinese_library')
# cursor = db.cursor()
# sql = 'select b_id from tb_library_book limit 10000;'
# cursor.execute(sql)
# data = cursor.fetchall()
# print(data)
# for i in range(10000):
#     update_sql = 'UPDATE tb_library_book SET b_listenlist_id = {0} WHERE b_id = {1};'.format(data2[i][0], "'"+str(data[i][0])+"'")
#     print(update_sql)
#     cursor.execute(update_sql)
#     db.commit()
#     # print(data[i][0])
# db.close()
# db2.close()


class MyMysql:

    def __init__(self,host, user, password, db):
        self.db = pymysql.connect(host=host, user=user,  password=password, db=db)
        self.cursor = self.db.cursor()

    def get_select(self, sql):
        self.cursor.execute(sql)
        data = self.cursor.fetchall()
        print('获取全部数据')
        return data

    def use_update(self, sql):
        self.cursor.execute(sql)
        self.db.commit()
        print('执行且提交成功')

    def put_excel(self, sql):
        a_z = ["A","B","C","D","E","F","G"]
        data = self.get_select(sql)
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(len(data)):
            for j in range(len(data[0])):
                doc = a_z[j] + str(i + 1)
                print(doc)
                ws[doc] = str(data[i][j])
        wb.save('temp_stu_uid_new.xlsx')
        print('已正常存储在excel中 另存为csv即可')

    def over_clean(self):
        self.db.close()

# mysql = MyMysql(host='rm-2ze9ycl21k1slu15zrw.mysql.rds.aliyuncs.com', user='english_radio_rw', password='english_radio_RW', db='english_radio')
mysql = MyMysql(host='rm-2ze524to8tfx5n794.mysql.rds.aliyuncs.com', user='db_write', password='ql7bo366DbqG', db='peiyou_coursereport')
result = mysql.put_excel('select student_id,student_uid,class_id,course_id,subject_id,lesson_num,learnreport_url from tb_stu_report_info GROUP BY student_uid, course_id LIMIT 5,10000;')
# print(result)

# num = 999910000
# for i in result:
#      num +=1
#      stuid = 'xzystudentid'+str(num)
#      mysql.use_update("UPDATE tb_student_card SET student_id ='{0}', student_uid ={1}  where id = '{2}';".format(stuid,num, i[0]))
# print(num)
#
# for i in range(10000):
#     num += 1
#     mysql.use_update(
#         "UPDATE tb_student_card SET student_id ='{0}', student_uid ={1}  where student_uid = '{2}';".format(stuid, num, i[0]))